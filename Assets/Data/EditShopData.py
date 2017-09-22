import sys
import os

import threading
import subprocess
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
import uuid

temp_uuid = str(uuid.uuid4()) + '.xlsx'

source_file = os.path.join(os.path.dirname(__file__), 'ShopData.json')
target_file = os.path.join(os.path.dirname(__file__), temp_uuid)

with open(source_file) as _f:
    data = json.load(_f)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'ShopData'

column_idx = 1
for buyable in data['_keys']:
    worksheet[get_column_letter(column_idx) + '1'] = buyable
    row_idx = 2
    if data['_values'][column_idx - 1]:
        for val in data['_values'][column_idx - 1].split(':'):
            worksheet[get_column_letter(column_idx) + str(row_idx)] = int(val)
            row_idx += 1
    column_idx += 1

workbook.save(target_file)


class AsyncReader(threading.Thread):

    def __init__(self):
        super(AsyncReader, self).__init__()
        self.change_date = os.path.getmtime(target_file)
        self.should_stop = False

    def read(self):
        if not os.path.exists(target_file):
            return
        if self.change_date == os.path.getmtime(target_file):
            return
        print('Updating!')
        self.change_date = os.path.getmtime(target_file)
        read_workbook = load_workbook(target_file)
        read_worksheet = read_workbook.get_active_sheet()

        res = {'_keys': [], '_values': []}

        buyable_column = 1
        while read_worksheet[get_column_letter(buyable_column) + '1'].value:
            res['_keys'].append(read_worksheet[get_column_letter(buyable_column) + '1'].value)
            print('key ' + res['_keys'][-1])
            row_index = 2
            rows = []
            while True:
                new_val = read_worksheet[get_column_letter(buyable_column) + str(row_index)].value
                try:
                    new_val = int(str(new_val))
                except ValueError:
                    break

                rows.append(str(new_val))
                print('val ' + str(len(rows)) + ' ' + str(rows[-1]))
                row_index += 1
            buyable_column += 1
            res['_values'].append(':'.join(rows))

        print(res)
        with open(source_file, 'w+') as _f:
            json.dump(res, _f)

    def run(self):
        while not self.should_stop:
            self.read()
            time.sleep(0.1)
        self.read()


class AsyncStartFile(threading.Thread):

    def __init__(self, filepath):
        super(AsyncStartFile, self).__init__()
        self.filepath = filepath

    def read(self):
        pass

    def run(self):
        print(self.filepath)
        while not os.path.exists(self.filepath):
            time.sleep(1)
        subprocess.Popen(['start', '/WAIT', self.filepath], shell=True).wait()
        print('Process finished')

read_thread = AsyncReader()
read_thread.start()

file_thread = AsyncStartFile(target_file)
file_thread.start()
file_thread.join()
read_thread.should_stop = True


os.remove(target_file)
input()



